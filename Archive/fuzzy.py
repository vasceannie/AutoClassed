import pandas as pd
from thefuzz import fuzz, process
from openpyxl import Workbook
from openpyxl.styles import Font
from concurrent.futures import ThreadPoolExecutor, as_completed
from tqdm import tqdm


# Infer matches using fuzzy matching with a 90% substring rule
def infer_matches(supplier_name, supplier_names, threshold=85):
    def match_condition(s1, s2):
        return (s1.lower() in s2.lower()) or (s2.lower() in s1.lower())

    matches = [
        (name, fuzz.token_sort_ratio(supplier_name, name))
        for name in supplier_names
        if match_condition(supplier_name, name)
    ]
    return [(match, score) for match, score in matches if score >= threshold]


# Process each row concurrently
def process_row(row, supplier_names):
    return infer_matches(row["Supplier Name"], supplier_names)


# Create hierarchical structure ensuring no duplicates and proper nesting
def create_hierarchical_structure(df):
    supplier_names_list = df["Supplier Name"].tolist()

    results = []
    with ThreadPoolExecutor(
        max_workers=36
    ) as executor:  # Estimate max workers for Intel i9-14900K
        futures = [
            executor.submit(process_row, row, supplier_names_list)
            for idx, row in df.iterrows()
        ]
        for future in tqdm(
            as_completed(futures), total=len(futures), desc="Processing Matches"
        ):
            results.append(future.result())

    supplier_matches = {idx: result for idx, result in enumerate(results) if result}
    hierarchy = []
    visited = set()

    for idx, row in df.iterrows():
        if idx in visited:
            continue
        matches = supplier_matches.get(idx, [])
        if matches:
            highest_spend_idx = max(
                [
                    next(iter(df.loc[df["Supplier Name"] == m[0]].index), idx)
                    for m in matches
                ],
                key=lambda x: df.at[x, "Spend"],
            )
        else:
            highest_spend_idx = idx

        total_spend = df.at[highest_spend_idx, "Spend"] + sum(
            df.at[match_idx, "Spend"]
            for match_idx in [
                next(iter(df.loc[df["Supplier Name"] == m[0]].index), idx)
                for m in matches
            ]
        )

        hierarchy.append((highest_spend_idx, idx, matches, total_spend))
        visited.update(
            next(iter(df.loc[df["Supplier Name"] == m[0]].index), idx) for m in matches
        )

    # Sort hierarchy by the number of matches (child rows) descending
    hierarchy.sort(key=lambda x: df.at[x[0], "Spend"], reverse=True)

    return hierarchy


# Write the hierarchical structure to an Excel file
def to_excel(df, hierarchy, output_path):
    wb = Workbook()
    ws = wb.active
    ws.title = "Hierarchical Suppliers"

    headers = ["Original Row Order"] + list(df.columns) + ["Parent Spend"]
    ws.append(headers)

    default_font = Font(bold=True)
    child_font = Font(bold=False, italic=True, color="505050")
    row_index = 2  # Starting from the second row since the first row is headers

    for parent_idx, original_idx, matches, total_spend in hierarchy:
        parent_row = df.loc[parent_idx]
        parent_row_formatted = (
            [parent_row["Original Row Order"]] + list(parent_row) + [total_spend]
        )
        ws.append(parent_row_formatted)
        for cell in ws[row_index]:
            cell.font = default_font
        row_index += 1

        for match, score in matches:
            match_idx = next(iter(df.loc[df["Supplier Name"] == match].index), None)
            if match_idx is None or match_idx == parent_idx:
                continue
            match_row = df.loc[match_idx]
            match_row_formatted = (
                [f'  {parent_row["Original Row Order"]}']
                + [
                    "  " + str(val) if isinstance(val, str) else val
                    for val in match_row
                ]
                + [None]
            )
            ws.append(match_row_formatted)
            for cell in ws[row_index]:
                cell.font = child_font
            row_index += 1

    wb.save(output_path)


# Main function to orchestrate the processing
def main(input_file_path, output_file_path):
    df = pd.read_excel(input_file_path, sheet_name="AP 2023 Suppliers")
    df["Original Row Order"] = range(1, len(df) + 1)

    hierarchy = create_hierarchical_structure(df)

    to_excel(df, hierarchy, output_file_path)
    print(f"Results saved to {output_file_path}")


if __name__ == "__main__":
    input_file_path = "data/ap_vendor_list.xlsx"  # Replace with your input file path
    output_file_path = "data/normalized_suppliers_output.xlsx"  # Replace with your desired output file path
    main(input_file_path, output_file_path)
    print(f"Results saved to {output_file_path}")
