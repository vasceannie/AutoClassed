import pandas as pd
from fuzzywuzzy import fuzz, process
from openpyxl import Workbook
from openpyxl.styles import Font
from tqdm import tqdm
from concurrent.futures import ThreadPoolExecutor, as_completed


def infer_matches(supplier_name, supplier_names, threshold=85):
    matches = process.extract(
        supplier_name, supplier_names, scorer=fuzz.token_sort_ratio
    )
    return [(match, score) for match, score in matches if score >= threshold]


def process_row(row, supplier_names):
    return infer_matches(row["Supplier Name"], supplier_names)


def create_hierarchical_structure(df):
    supplier_names_list = df["Supplier Name"].tolist()

    results = []
    with ThreadPoolExecutor(max_workers=10) as executor:
        futures = [
            executor.submit(process_row, row, supplier_names_list)
            for idx, row in df.iterrows()
        ]
        for future in tqdm(
            as_completed(futures), total=len(futures), desc="Processing Matches"
        ):
            results.append(future.result())

    supplier_matches = {idx: result for idx, result in enumerate(results)}

    # Determine parents and create hierarchical output
    hierarchy = []
    visited = set()

    for idx, row in df.iterrows():
        if idx in visited:
            continue
        matches = supplier_matches[idx]
        if matches:
            highest_spend_idx = max(
                [df.loc[df["Supplier Name"] == m[0]].index[0] for m in matches],
                key=lambda x: df.at[x, "Spend"],
            )
        else:
            highest_spend_idx = idx

        total_spend = df.at[highest_spend_idx, "Spend"] + sum(
            df.at[df.loc[df["Supplier Name"] == m[0]].index[0], "Spend"]
            for m in matches
        )

        hierarchy.append((highest_spend_idx, idx, matches, total_spend))
        visited.update(df.loc[df["Supplier Name"] == m[0]].index[0] for m in matches)

    return hierarchy


def to_excel(df, hierarchy, output_path):
    wb = Workbook()
    ws = wb.active
    ws.title = "Hierarchical Suppliers"

    # Write headers
    headers = list(df.columns) + ["Parent Spend"]
    ws.append(headers)

    # Write rows with hierarchical structure
    default_font = Font(bold=True)
    child_font = Font(bold=False, italic=True, color="505050")

    for parent_idx, original_idx, matches, total_spend in hierarchy:
        parent_row = df.loc[parent_idx]
        parent_row_formatted = list(parent_row) + [total_spend]
        ws.append(parent_row_formatted)
        for cell in ws[-1]:
            cell.font = default_font

        for match, score in matches:
            match_idx = df.loc[df["Supplier Name"] == match].index[0]
            if match_idx == parent_idx:
                continue
            match_row = df.loc[match_idx]
            match_row_formatted = [
                "  " + str(val) if isinstance(val, str) else val for val in match_row
            ] + [None]
            ws.append(match_row_formatted)
            for cell in ws[-1]:
                cell.font = child_font

    wb.save(output_path)


def main(input_file_path, output_file_path):
    df = pd.read_excel(input_file_path, sheet_name="AP 2023 Suppliers")
    df["Index"] = range(1, len(df) + 1)

    hierarchy = create_hierarchical_structure(df)

    to_excel(df, hierarchy, output_file_path)
    print(f"Results saved to {output_file_path}")


if __name__ == "__main__":
    input_file_path = "data/ap_vendor_list.xlsx"  # Replace with your input file path
    output_file_path = (
        "data/ap_vendor_list_cleaner.xlsx"  # Replace with your desired output file path
    )
    main(input_file_path, output_file_path)
    print(f"Results saved to {output_file_path}")
